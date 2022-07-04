from tempfile import NamedTemporaryFile
from barcode import EAN13
from barcode.writer import ImageWriter
import openpyxl as xl
from django.db import models
from django.core.files import File
from django.utils.translation import gettext_lazy as _
from webdev.materiais.models import Material

class Categoria(models.Model):
    nome = models.CharField(_('Categoria'), max_length=150, unique=True)

    def __str__(self):
        return self.nome

class Produto(models.Model):
    foto = models.ImageField(_('Foto do Produto'), upload_to='produtos', default='default.jpg', blank=True, null=True)
    nome = models.CharField(_('Nome'), max_length=150)
    colecao = models.CharField(_('Coleção'), max_length=150)
    familia = models.CharField(_('Família'), max_length=150, blank=True, null=True)
    observacao = models.TextField(_('Observação'), blank=True, null=True)
    categorias = models.ManyToManyField(Categoria, verbose_name=_('Categorias'), blank=True)
    data_criacao = models.DateField(_('Data de Criação'), blank=True, null=True)
    unidades = models.IntegerField(_('unidades em estoque'), default=0)
    tamanho = models.IntegerField(_('tamanho'), blank=True, null=True)
    barcode = models.ImageField(_('código de barras'), upload_to='produtos/barcode/', blank=True, null=True)

    def __str__(self):
        return f"{self.nome}"
    
    @classmethod
    def get_sortable_fields(cls):
        sortable_fields = [(f.name, f.verbose_name) for f in cls._meta.fields \
                            if f.name not in ['id', 'foto', 'observacao']]
        return sortable_fields

    def get_servicos(self):
        return ServicoDoProduto.objects.filter(produto=self)

    def get_materiais(self):
        return MaterialDoProduto.objects.filter(produto=self)

    def get_custo_de_producao(self):
        custo = 0
        for servico in self.get_servicos():
            custo += servico.valor
        for material_dp in self.get_materiais():
            custo += material_dp.get_custo()
        return round(custo, 2)

    def get_preco_atacado(self):
        return self.get_custo_de_producao() * 2

    def get_preco_revenda(self):
        return self.get_custo_de_producao() * 3

    def get_preco_cliente_final(self):
        return self.get_custo_de_producao() * 4
    
    def generate_barcode(self):
        img_temp_file = NamedTemporaryFile(delete=True)
        EAN13(format(self.id, '012'), writer=ImageWriter()).write(img_temp_file)
        temp_file = File(img_temp_file, name=f'{self.id}.png')
        if self.barcode:
            self.barcode.delete()
            self.save()
        self.barcode = temp_file
        self.save()
        return self.barcode
    
    def get_barcode(self):
        if self.barcode:
            return self.barcode
        else:
            return self.generate_barcode()

    def get_barcode_obj(self):
        return EAN13(format(self.id, '012'))

    def get_barcode_workbook(self):
        wb = xl.Workbook()
        ws = wb.active
        ws['A1'] = self.nome
        ws['A2'] = self.get_custo_de_producao()
        return wb


class MaterialDoProduto(models.Model):
    produto = models.ForeignKey(Produto, on_delete=models.CASCADE, verbose_name=_("produto"))
    material = models.ForeignKey(Material, on_delete=models.CASCADE, verbose_name=_("Material"))
    unidades = models.IntegerField(_("Unidades Utilizadas"), default=1)
    peso = models.DecimalField(_("Peso Unitário"), max_digits=8, decimal_places=2, blank=True, null=True) 
    UNIDADE_DE_MEDIDA_CHOICES = (
        ('', 'Unidade de Medida'),
        ('g', 'Gramas'),
        ('ct', 'Quilates'),
    )
    unidade_de_medida = models.CharField(_("Unidade de Medida"), max_length=2, choices=UNIDADE_DE_MEDIDA_CHOICES, blank=True, null=True)

    def __str__(self):
        return f"{self.material.nome} {self.unidades} unid."

    def get_peso(self):
        if self.peso and self.unidade_de_medida:
            return f"{self.peso} {self.unidade_de_medida}"
        else:
            return self.material.get_peso()

    def get_custo(self):
        if self.peso:
            preco = self.peso * self.material.get_preco_por_peso()
        else:
            preco= self.unidades * self.material.get_preco_unitario()
        return round(preco, 2)


class ServicoDoProduto(models.Model):
    produto = models.ForeignKey(Produto, on_delete=models.CASCADE, verbose_name=_("produto"))
    nome = models.CharField(_('nome'), max_length=150)
    valor = models.DecimalField(_("valor"), max_digits=8, decimal_places=2)

    def __str__(self):
        return self.nome
