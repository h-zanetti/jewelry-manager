import openpyxl
from tablib.core import Dataset
from django.db.models import Q
from django.shortcuts import render, redirect
from django.http import Http404, HttpResponseRedirect
from django.http.response import HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.forms import inlineformset_factory, modelformset_factory
from django.contrib import messages

from . import admin
from .models import Documento, Fornecedor, Fornecimento, Email, Telefone, Local, DadosBancarios, Servico
from .forms import FornecedorForm, FornecimentoForm, EmailForm, TelefoneForm, LocalForm, DadosBancariosForm, ServicoForm

@login_required
def meus_fornecedores(request):
    fornecedores = Fornecedor.objects.all()
    if request.GET:
        # Filter
        if 'search' in request.GET:
            fornecedores = fornecedores.filter(
                Q(nome__icontains=request.GET.get('search')) |
                Q(fornecimento__nome__icontains=request.GET.get('search'))
            )

    # Pagination
    paginator = Paginator(fornecedores, 10)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'title': 'Meus Fornecedores',
        'import_url': reverse('fornecedores:importar_fornecedores'),
        'export_url': reverse('fornecedores:exportar_fornecedores'),
        'create_url': reverse('fornecedores:novo_fornecedor'),
        'fornecedores': page_obj,
        # 'sort_form': sort_form,
        'sorting': True if 'sort-field' in request.GET else False,
        'sort_by': f'sort-field={request.GET.get("sort-field")}&sort-order={request.GET.get("sort-order")}',
        'search_by': f"search={request.GET.get('search')}" if 'search' in request.GET else None,
    }
    return render(request, 'fornecedores/meus_fornecedores.html', context)

@login_required
def novo_fornecedor(request):
    if request.method == 'POST':
        form = FornecedorForm(request.POST)
        if form.is_valid():
            fornecedor = form.save()
            next_url = request.POST.get('next')
            if next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:editar_fornecedor', fornecedor.id)
    else:
        form = FornecedorForm()

    context = {
        'title': 'Adicionar Novo Fornecedor',
        'form': form
    }

    return render(request, 'fornecedores/novo_fornecedor.html', context)

@login_required
def editar_fornecedor(request, fornecedor_id):
    try:
        fornecedor = Fornecedor.objects.get(id=fornecedor_id)
    except:
        raise Http404('Fornecedor não encontrado')

    EmailFormSet = inlineformset_factory(
        Fornecedor,
        Email,
        fields='__all__',
        extra=1
    )
    TelefoneFormSet = inlineformset_factory(
        Fornecedor,
        Telefone,
        fields='__all__',
        extra=1
    )
    LocalFormSet = inlineformset_factory(
        Fornecedor,
        Local,
        fields='__all__',
        extra=1
    )
    DadosBancariosFormSet = inlineformset_factory(
        Fornecedor,
        DadosBancarios,
        fields='__all__',
        extra=1
    )
    DocumentoFormSet = inlineformset_factory(
        Fornecedor,
        Documento,
        fields='__all__',
        extra=1
    )
    if request.method == 'POST':
        forms = {
            'fornecedor_form': FornecedorForm(request.POST, instance=fornecedor, prefix='fornecedor'),
            'email_formset': EmailFormSet(request.POST, instance=fornecedor, prefix='email'),
            'telefone_formset': TelefoneFormSet(request.POST, instance=fornecedor, prefix='telefone'),
            'local_formset': LocalFormSet(request.POST, instance=fornecedor, prefix='local'),
            'dados_bancarios_formset': DadosBancariosFormSet(request.POST, instance=fornecedor, prefix='dados_bancarios'),
            'documentos_formset': DocumentoFormSet(request.POST, instance=fornecedor, prefix='documento'),
        }
        for form in forms.values():
            if form.is_valid():
                form.save()
        return redirect('fornecedores:meus_fornecedores')
    else:
        forms = {
            'fornecedor': fornecedor,
            'fornecedor_form': FornecedorForm(instance=fornecedor, prefix='fornecedor'),
            'email_formset': EmailFormSet(instance=fornecedor, prefix='email'),
            'telefone_formset': TelefoneFormSet(instance=fornecedor, prefix='telefone'),
            'local_formset': LocalFormSet(instance=fornecedor, prefix='local'),
            'dados_bancarios_formset': DadosBancariosFormSet(instance=fornecedor, prefix='dados_bancarios'),
            'documentos_formset': DocumentoFormSet(instance=fornecedor, prefix='documento'),
        }

    context = forms
    context['title'] = f"Editar fornecedor - {fornecedor.nome}"

    return render(request, 'fornecedores/editar_fornecedor.html', context)

@login_required
def deletar_fornecedor(request, fornecedor_id):
    if request.method == 'POST':
        Fornecedor.objects.get(id=fornecedor_id).delete()
    return HttpResponseRedirect(reverse('fornecedores:meus_fornecedores'))

@login_required
def fornecimentos(request):
    FornecimentoFormSet = modelformset_factory(Fornecimento, fields='__all__', can_delete=True, form=FornecimentoForm)
    if request.method == 'POST':
        formset = FornecimentoFormSet(request.POST)
        if formset.is_valid():
            formset.save()
            next_url = request.POST.get('next')
            if 'submit-stay' in request.POST:
                return redirect('fornecedores:fornecimentos')
            elif next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        formset = FornecimentoFormSet()

    context = {
        'title': 'Fornecimentos Disponíveis',
        'formset': formset
    }

    return render(request, 'fornecedores/fornecimentos.html', context)

@login_required
def editar_fornecimento(request, fornecimento_id):
    try:
        fornecimento = Fornecimento.objects.get(id=fornecimento_id)
    except:
        raise Http404('Fornecimento não encontrado')

    if request.method == 'POST':
        form = FornecimentoForm(request.POST, instance=fornecimento)
        if form.is_valid():
            form.save()
            return redirect('fornecedores:meus_fornecedores')
    else:
        form = FornecimentoForm(instance=fornecimento)

    context = {
        'title': 'Editar fornecimento',
        'form': form,
        'novo_obj': False
    }

    return render(request, 'base_form_sm.html', context)

@login_required
def remover_fornecimento(request, fornecimento_id, fornecedor_id):
    if request.method == 'POST':
        fornecedor = Fornecedor.objects.get(id=fornecedor_id)
        f = Fornecimento.objects.get(id=fornecimento_id)
        fornecedor.fornecimento.remove(f)
        fornecedor.save()
    return HttpResponseRedirect(reverse('fornecedores:meus_fornecedores'))

@login_required
def novo_email(request, fornecedor_id):
    try:
        fornecedor = Fornecedor.objects.get(id=fornecedor_id)
    except:
        raise Http404('Fornecedor não encontrado')

    if request.method == 'POST':
        form = EmailForm(request.POST, initial={'fornecedor': fornecedor})
        if form.is_valid():
            form.save()
            next_url = request.POST.get('next')
            if 'submit-stay' in request.POST:
                return redirect('fornecedores:novo_email', kwargs={'fornecedor_id': fornecedor_id})
            elif next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        form = EmailForm(initial={'fornecedor': fornecedor})

    context = {
        'title': f'Adicionar novo email para {fornecedor.nome}',
        'form': form,
        'novo_obj': True
    }

    return render(request, 'base_form_sm.html', context)

@login_required
def editar_email(request, email_id):
    try:
        email = Email.objects.get(id=email_id)
    except:
        raise Http404('Email não encontrado')

    if request.method == 'POST':
        form = EmailForm(request.POST, instance=email)
        if form.is_valid():
            form.save()
            return redirect('fornecedores:meus_fornecedores')
    else:
        form = EmailForm(instance=email)

    context = {
        'title': 'Editar email',
        'form': form,
        'novo_obj': False
    }

    return render(request, 'base_form_sm.html', context)

@login_required
def deletar_email(request, email_id):
    if request.method == 'POST':
        Email.objects.get(id=email_id).delete()
    return HttpResponseRedirect(reverse('fornecedores:meus_fornecedores'))

@login_required
def novo_telefone(request, fornecedor_id):
    try:
        fornecedor = Fornecedor.objects.get(id=fornecedor_id)
    except:
        raise Http404('Fornecedor não encontrado')

    if request.method == 'POST':
        form = TelefoneForm(request.POST, initial={'fornecedor': fornecedor})
        if form.is_valid():
            form.save()
            next_url = request.POST.get('next')
            if 'submit-stay' in request.POST:
                return redirect('fornecedores:novo_telefone', kwargs={'fornecedor_id': fornecedor_id})
            elif next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        form = TelefoneForm(initial={'fornecedor': fornecedor})

    context = {
        'title': f'Adicionar novo número de Telefone para {fornecedor.nome}',
        'form': form,
        'novo_obj': True
    }

    return render(request, 'base_form_sm.html', context)

@login_required
def editar_telefone(request, telefone_id):
    try:
        telefone = Telefone.objects.get(id=telefone_id)
    except:
        raise Http404('Fornecimento não encontrado')

    if request.method == 'POST':
        form = TelefoneForm(request.POST, instance=telefone)
        if form.is_valid():
            form.save()
            return redirect('fornecedores:meus_fornecedores')
    else:
        form = TelefoneForm(instance=telefone)

    context = {
        'title': 'Editar telefone',
        'form': form,
        'novo_obj': False
    }

    return render(request, 'base_form_sm.html', context)

@login_required
def deletar_telefone(request, telefone_id):
    if request.method == 'POST':
        Telefone.objects.get(id=telefone_id).delete()
    return HttpResponseRedirect(reverse('fornecedores:meus_fornecedores'))

@login_required
def novo_local(request, fornecedor_id):
    try:
        fornecedor = Fornecedor.objects.get(id=fornecedor_id)
    except:
        raise Http404('Fornecedor não encontrado')

    if request.method == 'POST':
        form = LocalForm(request.POST, initial={'fornecedor': fornecedor})
        if form.is_valid():
            form.save()
            next_url = request.POST.get('next')
            if 'submit-stay' in request.POST:
                return redirect('fornecedores:novo_local', kwargs={'fornecedor_id': fornecedor_id})
            elif next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        form = LocalForm(initial={'fornecedor': fornecedor})

    context = {
        'title': f'Adicionar nova localização para {fornecedor.nome}',
        'form': form,
        'novo_obj': True
    }

    return render(request, 'base_form_md.html', context)

@login_required
def editar_local(request, local_id):
    try:
        local = Local.objects.get(id=local_id)
    except:
        raise Http404('Fornecimento não encontrado')

    if request.method == 'POST':
        form = LocalForm(request.POST, instance=local)
        if form.is_valid():
            form.save()
            return redirect('fornecedores:meus_fornecedores')
    else:
        form = LocalForm(instance=local)

    context = {
        'title': 'Editar localização',
        'form': form,
        'novo_obj': False
    }

    return render(request, 'base_form_md.html', context)

@login_required
def deletar_local(request, local_id):
    if request.method == 'POST':
        Local.objects.get(id=local_id).delete()
    return HttpResponseRedirect(reverse('fornecedores:meus_fornecedores'))

@login_required
def novos_dados_bancarios(request, fornecedor_id):
    try:
        fornecedor = Fornecedor.objects.get(id=fornecedor_id)
    except:
        raise Http404('Fornecedor não encontrado')

    if request.method == 'POST':
        form = DadosBancariosForm(request.POST, initial={'fornecedor': fornecedor})
        if form.is_valid():
            form.save()
            next_url = request.POST.get('next')
            if 'submit-stay' in request.POST:
                return redirect('fornecedores:novos_dados_bancarios', kwargs={'fornecedor_id': fornecedor_id})
            elif next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        form = DadosBancariosForm(initial={'fornecedor': fornecedor})

    context = {
        'title': f'Adicionar novos dados bancários para {fornecedor.nome}',
        'form': form,
        'novo_obj': True
    }

    return render(request, 'base_form_md.html', context)

@login_required
def editar_dados_bancarios(request, dados_bancarios_id):
    try:
        dados_bancarios = DadosBancarios.objects.get(id=dados_bancarios_id)
    except:
        raise Http404('Dados bancários não encontrados')

    if request.method == 'POST':
        form = DadosBancariosForm(request.POST, instance=dados_bancarios)
        if form.is_valid():
            form.save()
            return redirect('fornecedores:meus_fornecedores')
    else:
        form = DadosBancariosForm(instance=dados_bancarios)

    context = {
        'title': 'Editar dados bancários',
        'form': form,
        'novo_obj': False
    }

    return render(request, 'base_form_sm.html', context)

@login_required
def deletar_dados_bancarios(request, dados_bancarios_id):
    if request.method == 'POST':
        DadosBancarios.objects.get(id=dados_bancarios_id).delete()
    return HttpResponseRedirect(reverse('fornecedores:meus_fornecedores'))

@login_required
def novo_servico(request):
    if request.method == 'POST':
        form = ServicoForm(request.POST)
        if form.is_valid():
            form.save()
            next_url = request.POST.get('next')
            if 'submit-stay' in request.POST:
                return redirect('fornecedores:novo_servico')
            elif next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        form = ServicoForm()

    context = {
        'title': f'Cadastrar novo serviço',
        'form': form,
        'novo_obj': True
    }

    return render(request, 'fornecedores/servico_form.html', context)

@login_required
def editar_servico(request, servico_id):
    try:
        servico = Servico.objects.get(id=servico_id)
    except:
        raise Http404('Serviço não encontrado')

    if request.method == 'POST':
        form = ServicoForm(request.POST, instance=servico)
        if form.is_valid():
            form.save()
            next_url = request.POST.get('next')
            if next_url:
                return redirect(f'{next_url}')
            else:
                return redirect('fornecedores:meus_fornecedores')
    else:
        form = ServicoForm(instance=servico)

    context = {
        'title': 'Editar serviço',
        'form': form,
        'novo_obj': False
    }

    return render(request, 'fornecedores/servico_form.html', context)

@login_required
def deletar_servico(request, servico_id):
    if request.method == 'POST':
        Servico.objects.get(id=servico_id).delete()
        messages.success(request, 'Serviço deletado com sucesso.')
        next_url = request.POST.get('next')
        if next_url:
            return redirect(f'{next_url}')
        else:
            return redirect('fornecedores:meus_fornecedores')

# Importar e exportar fornecedores
@login_required
def exportar_fornecedores(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=fornecedores.xls'
    datasets = {
        'Fornecimentos': admin.FornecimentoResource().export(),
        'Fornecedores': admin.FornecedorResource().export(),
        'Emails': admin.EmailResource().export(),
        'Telefones': admin.TelefoneResource().export(),
        'Locais': admin.LocalResource().export(),
        'Dados_bancarios': admin.DadosBancariosResource().export(),
        'Documentos': admin.DocumentoResource().export(),
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.remove(ws)
    for name, data in datasets.items():
        ws = wb.create_sheet(name)
        ws.append(data._get_headers())
        for row in data._data:
            ws.append(list(row))
    wb.save(response)
    return response

@login_required
def importar_fornecedores(request):
    if request.method == 'POST':
        resources = {
            'Fornecimentos': admin.FornecimentoResource(),
            'Fornecedores': admin.FornecedorResource(),
            'Emails': admin.EmailResource(),
            'Telefones': admin.TelefoneResource(),
            'Locais': admin.LocalResource(),
            'Dados_bancarios': admin.DadosBancariosResource(),
            'Documentos': admin.DocumentoResource(),
        }
        datasets = {}
        for model in resources.keys():
            datasets[model] = Dataset()

        xl_file = request.FILES['myfile']
        wb = openpyxl.load_workbook(xl_file)
        for ws in wb.worksheets:
            data = datasets[ws.title]
            data.tite = ws.title
            for i, row in enumerate(ws.rows):
                row_vals = [c.value for c in row]
                if i==0:
                    data.headers = row_vals
                else:
                    data.append(row_vals)

        for name, resource in resources.items():
            resource.import_data(datasets[name])

        return redirect('fornecedores:meus_fornecedores')
        
    return render(request, 'base_form_file.html', {'title': "Importação de fornecedores"})
